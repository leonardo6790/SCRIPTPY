import collections
import hashlib
import os
import re
from pathlib import Path

import openpyxl


BASE = Path(r"C:\Users\david.bautista\Downloads\mayo")
PDF_ROOT = BASE / "FUENTE 43"
XLSX_ROOT = BASE / "plantillas por fuente" / "43"


def duplicated_count(values):
    return sum(count - 1 for count in collections.Counter(values).values() if count > 1)


def file_hash(path):
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def doc_key_from_name(path):
    match = re.search(r"\b(43)\s*-\s*(\d{1,10})\b", path.name, re.IGNORECASE)
    if not match:
        match = re.search(r"\b(43)\s*-\s*(\d{1,10})\b", str(path.parent), re.IGNORECASE)
    if not match:
        return None
    return match.group(1).zfill(2), match.group(2).zfill(10)


def validate_pdfs():
    print("PDFS")
    print("mes;pdfs;dup_numero;dup_hash")

    for month_dir in sorted([p for p in PDF_ROOT.iterdir() if p.is_dir()]):
        pdfs = sorted(month_dir.rglob("*.pdf"))
        keys = [doc_key_from_name(pdf) for pdf in pdfs]
        keys = [key for key in keys if key]
        hashes = [file_hash(pdf) for pdf in pdfs]
        print(
            f"{month_dir.name};{len(pdfs)};"
            f"{duplicated_count(keys)};{duplicated_count(hashes)}"
        )

        key_counter = collections.Counter(keys)
        repeated_keys = [key for key, count in key_counter.items() if count > 1]
        for key in repeated_keys[:20]:
            files = [str(pdf.relative_to(BASE)) for pdf in pdfs if doc_key_from_name(pdf) == key]
            print(f"  DUP_NUMERO {key}: {files}")


def validate_templates():
    print("PLANTILLAS")
    print("archivo;zdoc;ztrans;fuentes;dup_docs;dup_zdoc_rows;dup_ztrans_rows")

    for workbook_path in sorted(XLSX_ROOT.glob("*.xls*")):
        wb = openpyxl.load_workbook(workbook_path, data_only=True)

        if "ZDocument_Temp1" not in wb.sheetnames or "ZTransac_Temp1" not in wb.sheetnames:
            print(f"{workbook_path.name};SIN_HOJAS_ESPERADAS")
            continue

        zdocs = list(wb["ZDocument_Temp1"].iter_rows(min_row=2, values_only=True))
        ztrs = list(wb["ZTransac_Temp1"].iter_rows(min_row=2, values_only=True))
        fuentes = sorted({str(row[1]) for row in zdocs if row[1] not in (None, "")})
        dup_docs = duplicated_count((row[1], row[2]) for row in zdocs)
        dup_zdoc = duplicated_count(zdocs)
        dup_ztr = duplicated_count(ztrs)

        print(
            f"{workbook_path.name};{len(zdocs)};{len(ztrs)};{fuentes};"
            f"{dup_docs};{dup_zdoc};{dup_ztr}"
        )

        if dup_docs:
            counter = collections.Counter((row[1], row[2]) for row in zdocs)
            for key, count in list(counter.items()):
                if count <= 1:
                    continue
                print(f"  DUP_DOC {key}: {count}")


validate_pdfs()
validate_templates()
