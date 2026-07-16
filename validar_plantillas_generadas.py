import collections

import openpyxl


FILES = [
    ("55", "Abril", r"plantillas por fuente/PlantillaFuente55Abril.xlsx"),
    ("56", "Abril", r"plantillas por fuente/PlantillaFuente56Abril.xlsx"),
    ("55", "Mayo", r"plantillas por fuente/PlantillaFuente55Mayo.xlsx"),
    ("56", "Mayo", r"plantillas por fuente/PlantillaFuente56Mayo.xlsx"),
    ("55", "Junio", r"plantillas por fuente/PlantillaFuente55Junio.xlsx"),
    ("56", "Junio", r"plantillas por fuente/PlantillaFuente56Junio.xlsx"),
]


def duplicated_count(values):
    return sum(count - 1 for count in collections.Counter(values).values() if count > 1)


print("fuente;mes;zdoc;ztrans;fuentes_en_doc;dup_docs;dup_zdoc_rows;dup_ztrans_rows")

for fuente, mes, path in FILES:
    wb = openpyxl.load_workbook(path, data_only=True)
    zdocs = list(wb["ZDocument_Temp1"].iter_rows(min_row=2, values_only=True))
    ztrs = list(wb["ZTransac_Temp1"].iter_rows(min_row=2, values_only=True))
    fuentes = sorted({str(row[1]) for row in zdocs if row[1] not in (None, "")})
    dup_docs = duplicated_count((row[1], row[2]) for row in zdocs)
    dup_zdoc = duplicated_count(zdocs)
    dup_ztr = duplicated_count(ztrs)

    print(
        f"{fuente};{mes};{len(zdocs)};{len(ztrs)};{fuentes};"
        f"{dup_docs};{dup_zdoc};{dup_ztr}"
    )
